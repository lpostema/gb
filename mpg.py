import xml.etree.ElementTree as ET
import os, shutil
import datetime
import openpyxl
tree = ET.parse('Meeting Plan Generator.xml')
root = tree.getroot()

from PyPDF2.PyPDF2.pdf import PdfFileReader

while 1:
	today = datetime.date.today()
	daysToNextWednesday = datetime.timedelta((2 - datetime.date.weekday(today)) % 7)
	nextWednesday = today + daysToNextWednesday
	dateString = input("What date do you want to create a plan for? [YYYY-MM-DD format] (Press Enter for " + str(nextWednesday) + ")")
	if dateString == "":
		break
	try:
		 nextWednesday = datetime.datetime.strptime(dateString, '%Y-%m-%d')
	except ValueError:
		print("Invalid date. Please use YYYY-MM-DD format.")
		continue
	break

nextWednesdayMonth = str(nextWednesday.month) if (nextWednesday.month >= 10) else '0' + str(nextWednesday.month)
nextWednesdayDay = str(nextWednesday.day) if (nextWednesday.day >=10) else '0' + str(nextWednesday.day)
wednesdaysDate = str(nextWednesday.year) + '-' + nextWednesdayMonth + '-' + nextWednesdayDay

lessonFile = open('..\\2015-16 Gopher Buddies Bible Study.pdf', 'rb')
lessonPDF = PdfFileReader(lessonFile)
"""        TOCpage = lessonPDF.getPage(3)
        TOCstring = TOCpage.extractText()
        print(TOCstring)"""



shutil.copyfile('Meeting Planner.xlsx', wednesdaysDate + '\Meeting Planner.xlsx')

yearwb         = openpyxl.load_workbook('..\\2015 Lesson Plan.xlsx', data_only=True )
yearwb_formula = openpyxl.load_workbook('..\\2015 Lesson Plan.xlsx', data_only=False)
calendarws = yearwb.get_sheet_by_name("Calendar")
calendarws_formula = yearwb_formula.get_sheet_by_name("Calendar")
versesws = yearwb.get_sheet_by_name("Verses")
meetingDates = tuple(calendarws.iter_rows('A2:A39'))
for meetingDate in meetingDates:
    if meetingDate[0].value.date() == nextWednesday:
        break
meetingDate = meetingDate[0]
nextWednesdayRow = meetingDate.row
nextWednesdayLesson = calendarws[('C' + str(nextWednesdayRow))].value

#find lesson in pdf

for i in range(15, 363):
        pdfpage = lessonPDF.getPage(i)
        pdfstring = pdfpage.extractText()
        if nextWednesdayLesson in pdfstring:
                lessonStartPage = i+1
                break

#find range of pages
pdfstring = pdfstring[:pdfstring.rfind('\n')]  #string page number off
pdfLessonString = pdfstring[pdfstring.rfind('\n'):]

for i in range(lessonStartPage+1, 363):
        pdfpage = lessonPDF.getPage(i)
        pdfstring = pdfpage.extractText()
        if pdfLessonString not in pdfstring:
                lessonEndPage = i
                break

for plugin in root.findall('plugin'):
    for filename in plugin.findall('filelist'):
        for file in filename.iter():
            if 'name' in file.attrib.keys():
                startIndex = file.attrib['name'].find('change-date')
                if startIndex != -1:
                    newString = file.attrib['name'][:startIndex]
                    newString += wednesdaysDate
                    newString += '\\Meeting Planner.pdf'
                    file.set('name', newString)
            if 'pageselection' in file.attrib.keys():
                startIndex = file.attrib['pageselection'].find('change-page')
                if startIndex != -1:
                    newString = str(lessonStartPage) + '-' + str(lessonEndPage)
                    file.set('pageselection', newString)
    for destination in plugin.findall('destination'):
        #print(destination.tag, destination.attrib)
        if 'value' in destination.attrib.keys():
            startIndex = destination.attrib['value'].find('change-date')
            if startIndex != -1:
                newString = destination.attrib['value'][:startIndex]
                newString += wednesdaysDate
                newString += '\\Full Meeting Plan.pdf'
                destination.set('value', newString)
try:
    os.mkdir(wednesdaysDate)
    #print("Directory does not exist yet.")
except FileExistsError:
    print("Directory already exists.")
tree.write(wednesdaysDate + '\Meeting Plan Generator 2.xml')

        
plannerwb = openpyxl.load_workbook(wednesdaysDate + '\\Meeting Planner.xlsx')
meetingws = plannerwb.get_sheet_by_name('Meeting Overview')
meetingws['E10'] = nextWednesdayLesson
meetingws['G1'] = nextWednesday
# find theme
nextWednesdayThemeRow = nextWednesdayRow
nextWednesdayTheme = calendarws['D' + str(nextWednesdayThemeRow)].value
while(nextWednesdayTheme == None):
	nextWednesdayThemeRow -= 1
	nextWednesdayTheme = calendarws['D' + str(nextWednesdayThemeRow)].value

# put verse in cell 'c7'
verseRow = calendarws_formula['E' + str(nextWednesdayThemeRow)].value[(calendarws_formula['E' + str(nextWednesdayThemeRow)].value).index('B')+1:]
meetingws['C7'] = versesws['E' + verseRow].value
# put verse song in cell 'd8'
meetingws['D8'] = versesws['D' + verseRow].value
# make number format different for date
meetingws['G1'].number_format = '[$-409]mmmm\\ d\\,\\ yyyy;@'
lessonFile.close()
try:
	plannerwb.save(wednesdaysDate + '\\Meeting Planner2.xlsx')
except:
	print("Please close the meeting worksheet. ")
